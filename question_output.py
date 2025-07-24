from selenium import common, webdriver
By = webdriver.common.by.By # type: ignore
from time import sleep
from main import open_web, safe_click_button, wait_until_presence_of, click_elements_by_text
driver = open_web()
TAG_NAME = By.TAG_NAME
CLASS_NAME = By.CLASS_NAME
wait_until_presence_of(driver, CLASS_NAME, "semi-navigation-item-text")
driver.find_elements(CLASS_NAME, "semi-navigation-item-text")[1].click()
chapter_list = {}
# 打開網頁並執行一些操作
safe_click_button(driver, "選擇章節")
wait_until_presence_of(driver, TAG_NAME, 'body')
CSS_SELECTOR = By.CSS_SELECTOR
from itertools import count
from bs4 import BeautifulSoup, Tag
for i in count(1):
    # 定位到 semi-portal 裡的 ul 元素
    page_numbers = driver.find_elements(CLASS_NAME, 'semi-portal')[0].find_elements(CSS_SELECTOR, 'ul')[0].find_elements(TAG_NAME, 'li')
    for page_number in page_numbers:  # find_elements 如果在 in 裡面，好像找下個 page_number 時會再打一次？
        if page_number.text != str(i):
            continue
        # 點擊頁碼
        page_number.click()
        wait_until_presence_of(driver, TAG_NAME, 'body')
        # 提取當前頁面的章節資料
        # 解析網頁內容
        tables = BeautifulSoup(driver.page_source, 'html.parser').find_all('table')
        if len(tables) < 2:  # 如果沒有足夠的表格，返回空列表
            break
        second_table = tables[1].find_all('tr')
        for table_row in second_table:
            cells = table_row.find_all('td')
            if len(cells) >= 2:
                chapter_number = cells[1].text.strip()
                chapter_list[chapter_number] = f"{chapter_number} {cells[2].text.strip()}"
        break
    else:
        break
from tkinter import Tk, Frame, Checkbutton, StringVar, Button, LEFT, TOP, BOTTOM
chapter_vars = {}
selection_window = Tk()
selection_window.title("選擇章節")
selection_window.geometry("750x350")
pages = []
j = 0
frame = False
for chapter_number, chapter_title in chapter_list.items():
    if not frame:
        frame = Frame(selection_window)
        pages.append(frame)
    var = StringVar(value="0")
    Checkbutton(frame, text=chapter_title, variable=var).pack()
    j += 1
    chapter_vars[chapter_number] = var
    if j == 10:
        frame = False
        j = 0
page_buttons_frame = Frame(selection_window)
for i in range(len(pages)):
    def show_page(index):
        """頁碼按鈕"""
        for frame in pages:
            frame.pack_forget()
        pages[index].pack()
    Button(page_buttons_frame, text=str(i + 1), command=lambda index=i: show_page(index)).pack(side=LEFT)
page_buttons_frame.pack(side=TOP)
# 底部按鈕
bottom_buttons_frame = Frame(selection_window)
def submit_selections(download_choices):
    """確認按鈕"""
    selected_chapters = [chapter for chapter, var in chapter_vars.items() if var.get() == "1"]
    # Input chapters
    XPATH = By.XPATH
    input_box = driver.find_element(XPATH, '//*[@id="semi-modal-body"]/div/div[1]/div/input')
    input_box.click()
    for chapter in selected_chapters:
        input_box.send_keys(chapter)
        sleep(0.5)
        driver.find_element(XPATH, '//*[@id="semi-modal-body"]/div/div[2]/div/div/div/div[1]/div/table/tbody/tr[1]/td[1]/span/span/span/span').click()
        input_box.click()
        input_box.clear()
    driver.find_element(XPATH, '//*[@id="dialog-0"]/div/div[3]/div/button[2]/span').click()
    sleep(1)
    # 存儲所有頁面的資料
    columns = []
    data = []
    def pre_text_of(editor: Tag):
        return editor.find_all('pre')[1].text
    from json import dumps
    def isTag(tag) -> bool:
        return isinstance(tag, Tag)
    from backoff import expo, on_exception
    while True:
        texts = []
        for table_row in BeautifulSoup(driver.page_source, 'html.parser').find_all('tr'):
            if col_text := [col.text for col in table_row.find_all('td')]:
                texts.append(col_text)
            elif not columns:
                columns = [col.text for col in table_row.find_all('th')] + (["選項", "正確答案", "題解"] if download_choices else [])
        if download_choices:
            for i in range(len(rows := driver.find_elements(TAG_NAME, 'tr')[1:])):
                on_exception(wait_gen=expo, exception=common.exceptions.ElementClickInterceptedException)(
                    lambda: click_elements_by_text(rows[i], "semi-button-content", "修改"))()
                wait_until_presence_of(driver, By.ID, 'title-editor')
                page_source = BeautifulSoup(driver.page_source, 'html.parser')
                if isTag(semi_modal_body := page_source.find(id="semi-modal-body")):
                    page_source = semi_modal_body
                text_i = texts[i]
                find_source = page_source.find # type: ignore
                text_i[0] = pre_text_of(tag) if isTag(tag := find_source(id="title-editor")) else "" # type: ignore
                text_append = text_i.append
                text_append(dumps([pre_text_of(option_editor) for option_editor in page_source.select('div[id^=option-editor-]')])) # type: ignore
                if isTag(modelAnswer := find_source(attrs={'x-field-id': 'modelAnswer'})): # type: ignore
                    match text_i[1]:
                        case "單選題":
                            text_append(checked_answer.text if isTag(checked_answer := modelAnswer.find(class_='semi-radio-checked')) else "") # type: ignore
                        case "是非題":
                            text_append(isTag(modelAnswer.find(class_='semi-switch-checked'))) # type: ignore
                        case "複選題":
                            text_append(dumps([checkbox.text for checkbox in modelAnswer.find_all(class_='semi-checkbox-checked')])) # type: ignore
                        case _:
                            text_append("")
                else:
                    text_append("")
                text_append(pre_text_of(tag) if isTag(tag := find_source(id="note-editor")) else "") # type: ignore
                safe_click_button(driver, "取消")
        data += texts  # 將數據加到列表中
        next_button = driver.find_element(CLASS_NAME, 'semi-page-item.semi-page-next')  # 調整為實際的下一頁按鈕選擇器
        if next_button.get_attribute("aria-disabled") == "false":
            next_button.click()
            sleep(1)  # 等待頁面加載；如果真的要捉，需要捉頁碼是否改變，有點麻煩... 出錯時再改就好。
        else:
            break
    # 將資料轉換成 DataFrame
    from pandas import DataFrame, to_numeric
    df = DataFrame(data, columns=columns)  # 調整列名為實際情況
    # 計算答對比例
    df['練習次數'] = to_numeric(df['練習次數'], errors='coerce').fillna(0)
    df['答對次數'] = to_numeric(df['答對次數'], errors='coerce').fillna(0)
    df['答對比例'] = ((df['答對次數'] / df['練習次數']) * 100).round(2)  # 先計算百分比，再四捨五入到小數點後兩位

    # 將數值轉換為百分比格式的字符串
    df['答對比例'] = df['答對比例'].apply(lambda x: f'{x}%')

    # 將答對比例列移動到答對次數後面
    # 注意：如果你已經有 '答對比例' 列在 DataFrame 中，你可以透過列重排來達成，而不需要刪除再插入
    cols = df.columns.tolist()
    cols.insert(cols.index('答對次數') + 1, cols.pop(cols.index('答對比例')))
    df = df[cols]

    # 匯出到 Excel，保存到桌面
    from pathlib import Path
    base_path = Path(r'C:\Users\Public\Documents\StudyInIUB\Computer Science\career\國立中興大學\\線上測驗系統題庫')
    if len(selected_chapters) == 1 and base_path.exists():
        base_path /= chapter_list[selected_chapters[0]]
    from openpyxl import load_workbook, utils
    try:
        path = base_path.with_suffix('.xlsx')
        df.to_excel(path, index=False)

        # 然後，使用 openpyxl 加載剛剛保存的 Excel 檔案
        wb = load_workbook(path)
        ws = wb.active
        if ws is not None:
            # 調整每個欄位的寬度
            for column_cells in ws.columns:
                column = column_cells[0].column
                if column is not None:
                    ws.column_dimensions[utils.get_column_letter(column)].width = max(
                        min(127, 2 * len(str(cell.value))) for cell in column_cells)
        # 保存對 Excel 檔案所做的更改
        wb.save(path)
    except utils.exceptions.IllegalCharacterError as e: # type: ignore
        df.to_csv(base_path.with_suffix('.csv'), index=False)
    selection_window.destroy()
Button(bottom_buttons_frame, text="下載題目", command=lambda: submit_selections(False)).pack(side=LEFT)
Button(bottom_buttons_frame, text="下載題目及選項", command=lambda: submit_selections(True)).pack(side=LEFT)
Button(bottom_buttons_frame, text="取消", command=selection_window.destroy).pack(side=LEFT)
bottom_buttons_frame.pack(side=BOTTOM)
show_page(0)  # 顯示第一頁
selection_window.mainloop()
driver.quit()
